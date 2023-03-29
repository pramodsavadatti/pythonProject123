# This firmware belongs HERIN ELECTRONICS

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
from typing import Union
import tkinter as tk
from tkinter import *
import tkinter.font as font
from tkinter import Frame
from PIL import ImageTk, Image
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill  # Connect cell styles
from openpyxl.styles import Font, Fill  # Connect styles for text
from openpyxl import load_workbook
import datetime
import sys
import os
from PIL import Image, ImageTk
import serial
import json
import serial.tools.list_ports as port_lists
import tkinter
from PIL import Image, ImageDraw, ImageFont
from tkinter import messagebox
import time
# ws['A3'] = datetime.datetime.now()
from openpyxl.styles import colors  # Connect colors for text and cells
from serial import Serial


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.

def calibration():
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[10]
    ws1.title = "calibration"
    global v1c
    global v2c
    global v3c
    global v4c
    global v5c
    global v6c
    global v7c
    global v8c
    global v9c
    global v10c
    v1c = ws1['B2'].value
    v2c = ws1['C2'].value
    v3c = ws1['D2'].value
    v4c = ws1['E2'].value
    v5c = ws1['F2'].value
    v6c = ws1['G2'].value
    v7c = ws1['H2'].value
    v8c = ws1['I2'].value
    v9c = ws1['J2'].value
    v10c = ws1['K2'].value





def insert_excelparamter1(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[0]
    ws1.title = "Board1"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = "158704796973200"
    ws1['B' + str(k)] = barcodename
    ws1['C' + str(k)] = datetime.datetime.now()
    ws1['D' + str(k)] = listvoltage[1]
    ws1['E' + str(k)] = listvoltage[2]
    ws1['F' + str(k)] = listvoltage[3]
    ws1['G' + str(k)] = listvoltage[4]
    ws1['H' + str(k)] = listvoltage[5]
    ws1['I' + str(k)] = listvoltage[6]
    ws1['J' + str(k)] = listvoltage[7]
    ws1['L' + str(k)] = listvoltage[8]
    ws1['M' + str(k)] = listvoltage[9]
    ws1['N' + str(k)] = listvoltage[10]
    # 12v

    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.08, 1.35, 2.25, 0.675, 11.500,0.000,0.000, 0.000]
    maxvoltage = [5.50, 3.63, 1.32, 1.65, 2.75, 0.825,13.850,3.700, 3.700, 2.500]
    # refvoltage = [1.2,1.5,2.5,3.3,5,pot-3.3/1.78, 12,vcc-2.7/3.6, vccq-1.7/1.950]
    li = ["D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k), "I" + str(k), "J" + str(k)]
          # "L"+str(k),"M"+str(k),"N"+str(k)]
    j = 0
   # k=0
    for i in li:
        c = ws1[i]
        #k=k+1

        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            #if (k < 5):
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
            # top.after(0, lambda: messagebox.showerror("Test Result", c))

        if j < 7:
            ws1['K' + str(k)] = "FAIL"
            ws1['K' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)
        else:

            ws1['K' + str(k)] = "PASS"
            ws1['K' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)

    if j < 7:

        print("Test FAIL\n")

        labeltest = tkinter.Label(top, text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(
            row=3, column=9)
        next_button = Button(top, width=39, pady=7, text='Next',   border=0, state=DISABLED)
        next_button.place(x=1120, y=700)


    else:
        print("Test PASS\n")

        labeltest = tkinter.Label(top, text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(
            row=3, column=9)
        def next():

         step3()

        Button(top, width=39, pady=7, text='Next', bg='#57a1f8', fg='white', border=0, command=next).place(x=1120,
                                                                                                           y=700)



    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")


def insert_excelparamter2(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[0]
    ws1.title = "Board2"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = "158704796973200"
    ws1['B' + str(k)] = barcodename
    ws1['C' + str(k)] = datetime.datetime.now()
    ws1['D' + str(k)] = listvoltage[1]
    ws1['E' + str(k)] = listvoltage[2]
    ws1['F' + str(k)] = listvoltage[3]
    ws1['G' + str(k)] = listvoltage[4]
    ws1['H' + str(k)] = listvoltage[5]
    ws1['I' + str(k)] = listvoltage[6]
    ws1['J' + str(k)] = listvoltage[7]
    ws1['L' + str(k)] = listvoltage[8]
    ws1['M' + str(k)] = listvoltage[9]
    ws1['N' + str(k)] = listvoltage[10]
    # 12v

    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.08, 1.35, 2.25, 0.675, 11.500, 0.000, 0.000, 0.000]
    maxvoltage = [5.50, 3.63, 1.32, 1.65, 2.75, 0.825, 13.850, 3.700, 3.700, 2.500]
    # refvoltage = [1.2,1.5,2.5,3.3,5,pot-3.3/1.78, 12,vcc-2.7/3.6, vccq-1.7/1.950]
    li = ["D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k), "I" + str(k), "J" + str(k)]
    # "L"+str(k),"M"+str(k),"N"+str(k)]
    j = 0
    # k=0
    for i in li:
        c = ws1[i]
        # k=k+1

        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            # if (k < 5):
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
            # top.after(0, lambda: messagebox.showerror("Test Result", c))

        if j < 7:
            ws1['K' + str(k)] = "FAIL"
            ws1['K' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)
        else:

            ws1['K' + str(k)] = "PASS"
            ws1['K' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)

    if j < 7:

        print("Test FAIL\n")

        labeltest = tkinter.Label(top, text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(
            row=3, column=9)
        next_button = Button(top, width=39, pady=7, text='Next', border=0, state=DISABLED)
        next_button.place(x=1120, y=700)


    else:
        print("Test PASS\n")

        labeltest = tkinter.Label(top, text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(
            row=3, column=9)

        def next():

            step3()

        Button(top, width=39, pady=7, text='Next', bg='#57a1f8', fg='white', border=0, command=next).place(x=1120,
                                                                                                           y=700)

    # top.after(0, lambda: messagebox.showerror("Test Result", "Test Failed"))

    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")
def insert_excelparamter3(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[2]
    ws1.title = "Board3"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = barcodename
    ws1['B' + str(k)] = datetime.datetime.now()
    ws1['C' + str(k)] = listvoltage[1]
    ws1['D' + str(k)] = listvoltage[2]
    ws1['E' + str(k)] = listvoltage[3]
    ws1['F' + str(k)] = listvoltage[4]
    ws1['G' + str(k)] = listvoltage[5]
    ws1['H' + str(k)] = listvoltage[6]
    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.05, 1.35, 2.25, 0.65]
    maxvoltage = [5.25, 3.5, 1.3, 1.6, 2.65, 0.85]
    # refvoltage = [1.2,1.8,2.5,3.3,5]
    li = ["C" + str(k), "D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k)]
    j = 0
    for i in li:
        c = ws1[i]
        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
        if j == 6:
            ws1['I' + str(k)] = "PASS"
            ws1['I' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)
        else:
            ws1['I' + str(k)] = "FAIL"
            ws1['I' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)

    if j == 6:
        print("Test PASS\n")
        labeltest = tkinter.Label(text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(row=5,
                                                                                                                column=8)
    else:
        print("Test FAIL\n")
        labeltest = tkinter.Label(text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(row=5,
                                                                                                              column=8)
    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")


def insert_excelparamter4(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[3]
    ws1.title = "Board4"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = barcodename
    ws1['B' + str(k)] = datetime.datetime.now()
    ws1['C' + str(k)] = listvoltage[1]
    ws1['D' + str(k)] = listvoltage[2]
    ws1['E' + str(k)] = listvoltage[3]
    ws1['F' + str(k)] = listvoltage[4]
    ws1['G' + str(k)] = listvoltage[5]
    ws1['H' + str(k)] = listvoltage[6]
    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.05, 1.35, 2.25, 0.65]
    maxvoltage = [5.25, 3.5, 1.3, 1.6, 2.65, 0.85]
    # refvoltage = [1.2,1.8,2.5,3.3,5]
    li = ["C" + str(k), "D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k)]
    j = 0
    for i in li:
        c = ws1[i]
        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
        if j == 6:
            ws1['I' + str(k)] = "PASS"
            ws1['I' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)
        else:
            ws1['I' + str(k)] = "FAIL"
            ws1['I' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)

    if j == 6:
        print("Test PASS\n")
        labeltest = tkinter.Label(text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(row=6,
                                                                                                                column=8)
    else:
        print("Test FAIL\n")
        labeltest = tkinter.Label(text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(row=6,
                                                                                                              column=8)
    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")


def insert_excelparamter5(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[4]
    ws1.title = "Board5"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = barcodename
    ws1['B' + str(k)] = datetime.datetime.now()
    ws1['C' + str(k)] = listvoltage[1]
    ws1['D' + str(k)] = listvoltage[2]
    ws1['E' + str(k)] = listvoltage[3]
    ws1['F' + str(k)] = listvoltage[4]
    ws1['G' + str(k)] = listvoltage[5]
    ws1['H' + str(k)] = listvoltage[6]
    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.05, 1.35, 2.25, 0.65]
    maxvoltage = [5.25, 3.5, 1.3, 1.6, 2.65, 0.85]
    # refvoltage = [1.2,1.8,2.5,3.3,5]
    li = ["C" + str(k), "D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k)]
    j = 0
    for i in li:
        c = ws1[i]
        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
        if j == 6:
            ws1['I' + str(k)] = "PASS"
            ws1['I' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)
        else:
            ws1['I' + str(k)] = "FAIL"
            ws1['I' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)

    if j == 6:
        print("Test PASS\n")
        labeltest = tkinter.Label(text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(row=7,
                                                                                                                column=8)
    else:
        print("Test FAIL\n")
        labeltest = tkinter.Label(text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(row=7,
                                                                                                              column=8)
    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")


def insert_excelparamter6(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[5]
    ws1.title = "Board6"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = barcodename
    ws1['B' + str(k)] = datetime.datetime.now()
    ws1['C' + str(k)] = listvoltage[1]
    ws1['D' + str(k)] = listvoltage[2]
    ws1['E' + str(k)] = listvoltage[3]
    ws1['F' + str(k)] = listvoltage[4]
    ws1['G' + str(k)] = listvoltage[5]
    ws1['H' + str(k)] = listvoltage[6]
    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.05, 1.35, 2.25, 0.65]
    maxvoltage = [5.25, 3.5, 1.3, 1.6, 2.65, 0.85]
    # refvoltage = [1.2,1.8,2.5,3.3,5]
    li = ["C" + str(k), "D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k)]
    j = 0
    for i in li:
        c = ws1[i]
        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
        if j == 6:
            ws1['I' + str(k)] = "PASS"
            ws1['I' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)
        else:
            ws1['I' + str(k)] = "FAIL"
            ws1['I' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)

    if j == 6:
        print("Test PASS\n")
        labeltest = tkinter.Label(text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(row=8,
                                                                                                                column=8)
    else:
        print("Test FAIL\n")
        labeltest = tkinter.Label(text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(row=8,
                                                                                                              column=8)
    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")


def insert_excelparamter7(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[6]
    ws1.title = "Board7"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = barcodename
    ws1['B' + str(k)] = datetime.datetime.now()
    ws1['C' + str(k)] = listvoltage[1]
    ws1['D' + str(k)] = listvoltage[2]
    ws1['E' + str(k)] = listvoltage[3]
    ws1['F' + str(k)] = listvoltage[4]
    ws1['G' + str(k)] = listvoltage[5]
    ws1['H' + str(k)] = listvoltage[6]
    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.05, 1.35, 2.25, 0.65]
    maxvoltage = [5.25, 3.5, 1.3, 1.6, 2.65, 0.85]
    # refvoltage = [1.2,1.8,2.5,3.3,5]
    li = ["C" + str(k), "D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k)]
    j = 0
    for i in li:
        c = ws1[i]
        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
        if j == 6:
            ws1['I' + str(k)] = "PASS"
            ws1['I' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)
        else:
            ws1['I' + str(k)] = "FAIL"
            ws1['I' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)

    if j == 6:
        print("Test PASS\n")
        labeltest = tkinter.Label(text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(row=9,
                                                                                                                column=8)
    else:
        print("Test FAIL\n")
        labeltest = tkinter.Label(text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(row=9,
                                                                                                              column=8)
    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")


def insert_excelparamter8(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[7]
    ws1.title = "Board8"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = barcodename
    ws1['B' + str(k)] = datetime.datetime.now()
    ws1['C' + str(k)] = listvoltage[1]
    ws1['D' + str(k)] = listvoltage[2]
    ws1['E' + str(k)] = listvoltage[3]
    ws1['F' + str(k)] = listvoltage[4]
    ws1['G' + str(k)] = listvoltage[5]
    ws1['H' + str(k)] = listvoltage[6]
    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.05, 1.35, 2.25, 0.65]
    maxvoltage = [5.25, 3.5, 1.3, 1.6, 2.65, 0.85]
    # refvoltage = [1.2,1.8,2.5,3.3,5]
    li = ["C" + str(k), "D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k)]
    j = 0
    for i in li:
        c = ws1[i]
        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
        if j == 6:
            ws1['I' + str(k)] = "PASS"
            ws1['I' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)
        else:
            ws1['I' + str(k)] = "FAIL"
            ws1['I' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)

    if j == 6:
        print("Test PASS\n")
        labeltest = tkinter.Label(text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(row=10,
                                                                                                                column=8)
    else:
        print("Test FAIL\n")
        labeltest = tkinter.Label(text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(row=10,
                                                                                                              column=8)
    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")


def insert_excelparamter9(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[8]
    ws1.title = "Board9"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = barcodename
    ws1['B' + str(k)] = datetime.datetime.now()
    ws1['C' + str(k)] = listvoltage[1]
    ws1['D' + str(k)] = listvoltage[2]
    ws1['E' + str(k)] = listvoltage[3]
    ws1['F' + str(k)] = listvoltage[4]
    ws1['G' + str(k)] = listvoltage[5]
    ws1['H' + str(k)] = listvoltage[6]
    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.05, 1.35, 2.25, 0.65]
    maxvoltage = [5.25, 3.5, 1.3, 1.6, 2.65, 0.85]
    # refvoltage = [1.2,1.8,2.5,3.3,5]
    li = ["C" + str(k), "D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k)]
    j = 0
    for i in li:
        c = ws1[i]
        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
        if j == 6:
            ws1['I' + str(k)] = "PASS"
            ws1['I' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)
        else:
            ws1['I' + str(k)] = "FAIL"
            ws1['I' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)

    if j == 6:
        print("Test PASS\n")
        labeltest = tkinter.Label(text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(row=11,
                                                                                                                column=8)
    else:
        print("Test FAIL\n")
        labeltest = tkinter.Label(text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(row=11,
                                                                                                              column=8)
    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")


def insert_excelparamter10(listvoltage, barcodename):
    wb = Workbook()
    # wb.save("testresult.xlsx")
    wb = load_workbook('testresult.xlsx')
    ws1 = wb.worksheets[9]
    ws1.title = "Board10"

    # ws2 =wb.active
    k = str(ws1._current_row)
    k = int(k) + 1
    ws1._current_row = k + 1
    ws1['A' + str(k)] = barcodename
    ws1['B' + str(k)] = datetime.datetime.now()
    ws1['C' + str(k)] = listvoltage[1]
    ws1['D' + str(k)] = listvoltage[2]
    ws1['E' + str(k)] = listvoltage[3]
    ws1['F' + str(k)] = listvoltage[4]
    ws1['G' + str(k)] = listvoltage[5]
    ws1['H' + str(k)] = listvoltage[6]
    # ws.cell(int(k)+1,1)
    minvoltage = [4.5, 2.97, 1.05, 1.35, 2.25, 0.65]
    maxvoltage = [5.25, 3.5, 1.3, 1.6, 2.65, 0.85]
    # refvoltage = [1.2,1.8,2.5,3.3,5]
    li = ["C" + str(k), "D" + str(k), "E" + str(k), "F" + str(k), "G" + str(k), "H" + str(k)]
    j = 0
    for i in li:
        c = ws1[i]
        # c.font = Font(size=23,color='FF0000', bold=True, italic=True)
        if minvoltage[li.index(i)] < listvoltage[li.index(i) + 1] < maxvoltage[li.index(i)]:
            c.font = Font(size=12, color='00FF00', bold=True, italic=True)
            j = j + 1
        else:
            c.font = Font(size=12, color='FF0000', bold=True, italic=True)
        if j == 6:
            ws1['I' + str(k)] = "PASS"
            ws1['I' + str(k)].font = Font(size=12, color='00FF00', bold=True, italic=True)
        else:
            ws1['I' + str(k)] = "FAIL"
            ws1['I' + str(k)].font = Font(size=12, color='FF0000', bold=True, italic=True)

    if j == 6:
        print("Test PASS\n")
        labeltest = tkinter.Label(text="Pass", width=8, height=1, bg="white", fg="green", padx=10, pady=5).grid(row=12,
                                                                                                                column=8)
    else:
        print("Test FAIL\n")
        labeltest = tkinter.Label(text="Fail", width=8, height=1, bg="white", fg="red", padx=10, pady=5).grid(row=12,
                                                                                                              column=8)
    # import datetime
    # ws['A3'] = datetime.datetime.now()

    wb.save("testresult.xlsx")


def get_data():
    # print(listvoltage)
    # ser.write(b'q')
    flag = 0
    # time.sleep(0.2)
    command_list = [b'158704796973200']
    # , b'2', b'3', b'4', b'5', b'6', b'7', b'8', b'9', b'A']
    for c_list in command_list:
        i = 0
        # ser.write(b'q')
        # ser.write(b'q')
        # ser.write(b'q')
        # ser.flushInput()
        # ser.write(b'q')
        # ser.write(b'q')
        # ser.write(b'q')
        ser.write(c_list)
        line = ser.readline()

        # line = "{\"board\":\"3.0\",\"v1\":\"1.2\",\"v2\":\"1.8\",\"v3\":\"2.5\",\"v4\":\"3.3\",\"v5\":\"5.0\",\"v6\":\"5.0\"}"
        print(line)
        ser.flushInput()
        # top.after(0, lambda: messagebox.showerror("Test Result", line))
        # ser.write(b'q')
        # # time.sleep(0.2)
        # ser.write(b'q')
        # ser.write(b'q')
        # ser.write(b'q')
        flag = 0
        try:
            jsonobject = json.loads(line)
        # print(line)
        except ValueError as err:
            print("invalid jsondata")
            flag = 1
        # ser.write(b'q')
        # ser.flushInput()
        if flag == 0:
            calibration()
            inputvoltage = []
            inputvoltage.append((float(jsonobject["board"])))
            inputvoltage.append(round(float(jsonobject["v1"])+float(v1c),3))
            # 5v
            inputvoltage.append(round(float(jsonobject["v2"])+float(v2c),3))
            # 3.3v
            inputvoltage.append(round(float(jsonobject["v3"])+float(v3c),3))
            # 2.5v
            inputvoltage.append(round(float(jsonobject["v4"])+float(v4c),3))
            # 1.5v
            inputvoltage.append(round(float(jsonobject["v5"])+float(v5c),3))
            # 1.2v
            inputvoltage.append(round(float(jsonobject["v6"])+float(v6c),3))
            # 0.78v
            inputvoltage.append(round(float(jsonobject["v7"])+float(v7c),3))
            # P1.8/3/3v
            inputvoltage.append(round(float(jsonobject["v8"])+float(v8c),3))
            # 12v
            inputvoltage.append(round(float(jsonobject["v9"])+float(v9c),3))
            # # Vcc
            inputvoltage.append(round(float(jsonobject["v10"])+float(v10c),3))

            # Vccq
            # 12v
            # inputvoltage = [1.20,1.80,2.50,3.62,5.00]
            # print(inputvoltage)
            # get_data((inputvoltage))
            # j=0
            listvoltage = inputvoltage
            k1 = 0
            k2 = 0
            k3 = 0
            k4 = 0
            k5 = 0
            k6 = 0
            k7 = 0
            k8 = 0
            k9 = 0
            k10 = 0
            k11 = 0
            k12 = 0
            minvoltage = [4.5, 2.97, 1.08, 1.35, 2.25, 0.675, 11.500, 0.000, 0.000, 0.000]
            maxvoltage = [5.50, 3.63, 1.32, 1.65, 2.75, 0.825, 13.850, 3.700, 3.700, 2.500]
            print(len(listvoltage))
            i=0
            print(listvoltage)

            failed_voltages = []

            for lista in listvoltage:
                if (listvoltage[0] == 158704796973200.0 and listvoltage[0] != lista):
                    if (minvoltage[i] < lista < maxvoltage[i]):
                        #print(listvoltage.index(lista))
                        if (i<6):
                             tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="green", padx=8, pady=5).grid(
                                row=3, column=i + 2, padx=5, pady=5)

                        elif (i == 7):
                            tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="green", padx=8,
                                          pady=5).grid(
                                row=3, column=10, padx=5, pady=5)

                        elif (i == 6):
                            tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="green", padx=8,
                                          pady=5).grid(
                                row=3, column=8, padx=5, pady=5)
                        elif (i == 8):
                            tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="green", padx=8,
                                          pady=5).grid(
                                row=3, column=11, padx=5, pady=5)
                        elif (i == 9):
                            tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="green", padx=8,
                                          pady=5).grid(
                                row=3, column=12, padx=5, pady=5)
                    else:

                        if (i <= 6):
                            tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="red", padx=8,
                                          pady=5).grid(
                                row=3, column=i + 2, padx=5, pady=5)
                            failed_voltages.append((lista, f"Failed V {i}-Read the above descriptions "))
                        elif (i == 7):
                            tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="red", padx=8,
                                          pady=5).grid(
                                row=3, column=10, padx=5, pady=5)

                        elif (i == 6):
                            tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="red", padx=8,
                                          pady=5).grid(
                                row=3, column=8, padx=5, pady=5)
                        elif (i == 8):
                            tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="red", padx=8,
                                          pady=5).grid(
                                row=3, column=11, padx=5, pady=5)
                        elif (i == 9):
                            tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="red", padx=8,
                                          pady=5).grid(
                                row=3, column=12, padx=5, pady=5)



                    i = i + 1
                    if k1 == 0:
                        k1 = 1
                        insert_excelparamter1(listvoltage, entry1_var.get())

            if failed_voltages:
                # create a new window to display the failed voltages
                failed_window = tkinter.Toplevel(top)
                failed_window.title("Failed Voltages")
                failed_window.resizable(False, False)
                failed_window.geometry('925x500+300+200')
                # create a label to display the failed voltages
                failed_label = tkinter.Label(failed_window, text="The following voltages failed: If V6:(12V) is failed Check Fuse (or) On Off Controls          "
                                                                 "Else:Replace the Board", font=('Microsoft yaHei UI Light', 14, 'bold'), wraplength=300)
                failed_label.pack()
                # iterate over the failed voltages and display the voltage value and description text
                for voltage, description in failed_voltages:

                    failed_voltage_label = tkinter.Label(failed_window, text=f"{voltage}: {description} ",font=('Microsoft yaHei UI Light', 12, 'bold'),fg="red")
                    failed_voltage_label.pack()

                # create labels for each failed voltage

                if (listvoltage[0] == 2.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(row=4,
                                                                                                                  column=i + 2,
                                                                                                                  padx=5,
                                                                                                                  pady=5)


                    i = i + 1
                    if k2 == 0:
                        k2 = 1
                        insert_excelparamter2(listvoltage, entry1_var.get())


                if (listvoltage[0] == 3.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(row=5,
                                                                                                                  column=i + 2,
                                                                                                                  padx=5,
                                                                                                                  pady=5)
                    i = i + 1
                    # print(entry3_var.get())
                    if k3 == 0:
                        k3 = 1
                        insert_excelparamter3(listvoltage, entry3_var.get())
                if (listvoltage[0] == 4.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(row=6,
                                                                                                                  column=i + 2,
                                                                                                                  padx=5,
                                                                                                                  pady=5)
                    i = i + 1
                    # print(entry4_var.get())
                    if k4 == 0:
                        k4 = 1
                        insert_excelparamter4(listvoltage, entry4_var.get())
                if (listvoltage[0] == 5.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(row=7,
                                                                                                                  column=i + 2,
                                                                                                                  padx=5,
                                                                                                                  pady=5)
                    i = i + 1
                    # print(entry5_var.get())
                    if k5 == 0:
                        k5 = 1
                        insert_excelparamter5(listvoltage, entry5_var.get())
                if (listvoltage[0] == 6.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(row=8,
                                                                                                                  column=i + 2,
                                                                                                                  padx=5,
                                                                                                                  pady=5)
                    i = i + 1
                    # print(entry6_var.get())
                    if k6 == 0:
                        k6 = 1
                        insert_excelparamter6(listvoltage, entry6_var.get())
                if (listvoltage[0] == 7.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(row=9,
                                                                                                                  column=i + 2,
                                                                                                                  padx=5,
                                                                                                                  pady=5)
                    i = i + 1
                    # print(entry7_var.get())
                    if k7 == 0:
                        k7 = 1
                        insert_excelparamter7(listvoltage, entry7_var.get())
                if (listvoltage[0] == 8.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(
                        row=10, column=i + 2, padx=5, pady=5)
                    i = i + 1
                    # print(entry8_var.get())
                    if k8 == 0:
                        k8 = 1
                        insert_excelparamter8(listvoltage, entry8_var.get())
                if (listvoltage[0] == 9.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(
                        row=11, column=i + 2, padx=5, pady=5)
                    i = i + 1
                    # print(entry9_var.get())
                    if k9 == 0:
                        k9 = 1
                        insert_excelparamter9(listvoltage, entry9_var.get())
                if (listvoltage[0] == 10.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(
                        row=12, column=i + 2, padx=5, pady=5)
                    i = i + 1
                    # print(entry10_var.get())
                    if k10 == 0:
                        k10 = 1
                        insert_excelparamter10(listvoltage, entry10_var.get())
                if (listvoltage[0] == 11.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(
                        row=13, column=i + 2, padx=5, pady=5)
                    i = i + 1
                    # print(entry11_var.get())
                    if k11 == 0:
                        k11 = 1
                if (listvoltage[0] == 12.0 and listvoltage[0] != lista):
                    tkinter.Label(top, text=lista, width=8, height=1, bg="white", fg="blue", padx=8, pady=5).grid(
                        row=14, column=i + 2, padx=5, pady=5)
                    i = i + 1
                    # Sprint(entry12_var.get())

            ser.flushInput()
            # time.sleep(1)


global ports
ports = []


def dashbord():

    # print_hi('PyCharm')
    print("welcomes to micron technologies")
    print("1. Press 1 for testing\n")
    print("2. Press 2 to exit\n")

    def create_gradient():
        gradient = Image.new('RGB', (1920, 1080), (255, 255, 255))
        for x in range(1920):
            r = int(176* (1 - x / 1920))
            g = int(196* (1 - x / 1920))
            b = int(222 * (1 - x / 1920))
            for y in range(1080):
                gradient.putpixel((x, y), (r, g, b,))
        return gradient

    def select_com():
        global ser
        # print(vm)
        vm = clicked.get()
        for nim in ports:
            # print(nim)
            if (str(vm) == str(nim)):
                print("hi")
                va = nim.name
                print(va)
        ser = serial.Serial(va, 115200, timeout=1)
        print("connected")
        # select.configure(text="connected")
        labelcom = tkinter.Label(top, text="Connected", width=10, height=1, bg="Green", fg="White", padx=10,
                                 pady=5).grid(row=0, column=12)


    global top
    top = Toplevel(root1)
    global clicked
    clicked = tkinter.StringVar(top)
    #top.config(bg="grey")
    top.geometry("1920x1080")
    gradient = create_gradient()
    gradient = ImageTk.PhotoImage(gradient)
    label = Label(top, image=gradient)
    label.place(x=0, y=0, relwidth=1, relheight=1)

    top.title('Welcome to Herin Electronics')

    logo1 = ImageTk.PhotoImage(Image.open('micron_logo.png'))
    logo_label = tkinter.Label(top, image=logo1).grid(column=0, row=0, padx=5, )
    logo = ImageTk.PhotoImage(Image.open('logo .jpg'))
    logo_label = tkinter.Label(top, image=logo).grid(column=1, row=0, padx=5, )

    top.option_add("*font", "aerial 10 bold")
    l = tkinter.Label(top, text="Metha Test Jig Board", width=18, height=1, bg="light slate blue", fg="Blue", padx=8,
                      pady=5).grid(row=0, columnspan=3, column=2)
    canvas = tkinter.Canvas(top, width=1920, height=5, bg="white", highlightthickness=0, )
    # canvas.create_line(0,2,1920,2,width=5,fill="blue")
    canvas.grid(row=1, columnspan=100, pady=10)

    top.option_add("*font", "lucida 8 bold ")
    b = tkinter.Button(top, text="Getdata", width=8, height=1, bg="light slate blue", fg="Blue", padx=10, pady=5,
                       command=get_data).grid(row=0, column=5)
    label1 = tkinter.Label(top, text="MAC ID", width=12, height=1, bg="light slate blue", fg="Blue", padx=10,
                           pady=5).grid(row=2, column=0)

    vscan = tkinter.Button(top, text="Metha SN", width=10, height=1, bg="orange", fg="white", padx=10, pady=5).grid(
        row=2, column=1, padx=10, pady=5)

    v5 = tkinter.Button(top, text="5V", width=5, height=1, bg="white", fg="Black", padx=10, pady=5).grid(row=2,
                                                                                                         column=2,
                                                                                                         padx=10,
                                                                                                         pady=5)
    v3_3 = tkinter.Button(top, text="3.3V", width=6, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                             column=3,
                                                                                                             padx=10,
                                                                                                             pady=5)
    v1_2 = tkinter.Button(top, text="1.2V", width=6, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                             column=4,
                                                                                                             padx=10,
                                                                                                             pady=5)
    v1_5 = tkinter.Button(top, text="1.5V", width=6, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                             column=5,
                                                                                                             padx=10,
                                                                                                             pady=5)
    v2_5 = tkinter.Button(top, text="2.5V", width=6, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                             column=6,
                                                                                                             padx=10,
                                                                                                             pady=5)
    v0_78 = tkinter.Button(top, text="0.78V", width=7, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                               column=7,
                                                                                                               padx=10,
                                                                                                               pady=5)
    POT = tkinter.Button(top, text="POT", width=6, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                           column=10,
                                                                                                           padx=10,
                                                                                                           pady=5)
    v12 = tkinter.Button(top, text="F1-12V", width=6, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                           column=8,
                                                                                                           padx=10,
                                                                                                           pady=5)


    result = tkinter.Button(top, text="Results", width=8, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                                  column=9,
                                                                                                                  padx=10,
                                                                                                                  pady=5)
    SP1 = tkinter.Button(top, text="SP1", width=8, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                                  column=11,
                                                                                                                  padx=10,
                                                                                                                  pady=5)
    SP2= tkinter.Button(top, text="SP2", width=8, height=1, bg="white", fg="black", padx=10, pady=5).grid(row=2,
                                                                                                                  column=12,
                                                                                                                  padx=10,
                                                                                                                  pady=5)

    grid_frame = Frame(top)
    for row in range(3, 15):
        for column in range(2, 13):
            label = Label(top, width=8, height=1, bg="white", fg="blue", padx=8, pady=5)
            label.grid(row=row, column=column, padx=5, pady=5)
            grid_frame.grid_columnconfigure(column, weight=1)

    # Entry(top, width=10, borderwidth=5).grid(row=3, column=1)
    # Entry.grid(row=3,column=1, padx=10,pady=10)

    ports = list(port_lists.comports())
    for p in ports:
        print(p)
    # serin = input("Enter COM port name")
    print(ports[0])
    clicked.set(ports[0].name)
    # clicked.trace('w',lamda *args:Option_selected, com_connect)
    dropdown1 = tkinter.OptionMenu(top, clicked, *ports).grid(row=0, columnspan=3, column=8, padx=0, pady=0)
    connect = tkinter.Button(top, text="connect", command=select_com, width=8, height=1, bg="orange", fg="black",
                             padx=10, pady=5).grid(row=0, column=11, padx=10, pady=5)

    global entry1_var
    global entry2_var
    global entry3_var
    global entry4_var
    global entry5_var
    global entry6_var
    global entry7_var
    global entry8_var
    global entry9_var
    global entry10_var
    entry1_var = tkinter.StringVar()
    entry2_var = tkinter.StringVar()
    entry3_var = tkinter.StringVar()
    entry4_var = tkinter.StringVar()
    entry5_var = tkinter.StringVar()
    entry6_var = tkinter.StringVar()
    entry7_var = tkinter.StringVar()
    entry8_var = tkinter.StringVar()
    entry9_var = tkinter.StringVar()
    entry10_var = tkinter.StringVar()
    entry11_var = tkinter.StringVar()
    entry12_var = tkinter.StringVar()
    macid1_var = tkinter.StringVar()
    macid2_var = tkinter.StringVar()
    macid3_var = tkinter.StringVar()
    macid4_var = tkinter.StringVar()
    macid5_var = tkinter.StringVar()
    macid6_var = tkinter.StringVar()
    macid7_var = tkinter.StringVar()
    macid8_var = tkinter.StringVar()
    macid9_var = tkinter.StringVar()
    macid10_var = tkinter.StringVar()
    macid11_var = tkinter.StringVar()
    macid12_var = tkinter.StringVar()
    workbook = openpyxl.load_workbook('testresult.xlsx')

    # Select the worksheet where the data is located
    worksheet = workbook['MACID']

    # Get the values from column B and assign them to the StringVar() objects
    macid2_var.set(worksheet['B3'].value)
    macid3_var.set(worksheet['B4'].value)
    macid4_var.set(worksheet['B5'].value)
    macid5_var.set(worksheet['B6'].value)
    macid6_var.set(worksheet['B7'].value)
    macid7_var.set(worksheet['B8'].value)
    macid8_var.set(worksheet['B9'].value)
    macid9_var.set(worksheet['B10'].value)
    macid10_var.set(worksheet['B11'].value)
    macid11_var.set(worksheet['B12'].value)
    macid12_var.set(worksheet['B13'].value)

    # Close the workbook
    workbook.close()

    entry1 = tkinter.Entry(top, textvariable=entry1_var, width=10, borderwidth=5).grid(row=3, column=1, padx=10,
                                                                                       pady=10)
    entry2 = tkinter.Entry(top, textvariable=entry2_var, width=10, borderwidth=5).grid(row=4, column=1, padx=10,
                                                                                       pady=10)
    entry3 = tkinter.Entry(top, textvariable=entry3_var, width=10, borderwidth=5).grid(row=5, column=1, padx=10,
                                                                                       pady=10)
    entry4 = tkinter.Entry(top, textvariable=entry4_var, width=10, borderwidth=5).grid(row=6, column=1, padx=10,
                                                                                       pady=10)
    entry5 = tkinter.Entry(top, textvariable=entry5_var, width=10, borderwidth=5).grid(row=7, column=1, padx=10,
                                                                                       pady=10)
    entry6 = tkinter.Entry(top, textvariable=entry6_var, width=10, borderwidth=5).grid(row=8, column=1, padx=10,
                                                                                       pady=10)
    entry7 = tkinter.Entry(top, textvariable=entry7_var, width=10, borderwidth=5).grid(row=9, column=1, padx=10,
                                                                                       pady=10)
    entry8 = tkinter.Entry(top, textvariable=entry8_var, width=10, borderwidth=5).grid(row=10, column=1, padx=10,
                                                                                       pady=10)
    entry9 = tkinter.Entry(top, textvariable=entry9_var, width=10, borderwidth=5).grid(row=11, column=1, padx=10,
                                                                                       pady=10)
    entry10 = tkinter.Entry(top, textvariable=entry10_var, width=10, borderwidth=5).grid(row=12, column=1, padx=10,
                                                                                         pady=10)
    entry11 = tkinter.Entry(top, textvariable=entry10_var, width=10, borderwidth=5).grid(row=13, column=1, padx=10,
                                                                                         pady=10)
    entry12 = tkinter.Entry(top, textvariable=entry10_var, width=10, borderwidth=5).grid(row=14, column=1, padx=10,
                                                                                         pady=10)
    macid1 = tkinter.Label(top, text="158704796973200", width=15, borderwidth=5).grid(row=3, column=0, padx=10,
                                                                                      pady=10)
    macid2 = tkinter.Label(top, textvariable=macid2_var, width=15, borderwidth=5).grid(row=4, column=0, padx=10,
                                                                                       pady=10)
    macid3 = tkinter.Label(top, textvariable=macid3_var, width=15, borderwidth=5).grid(row=5, column=0, padx=10,
                                                                                       pady=10)
    macid4 =  tkinter.Label(top, textvariable=macid4_var, width=15, borderwidth=5).grid(row=6, column=0, padx=10,
                                                                                       pady=10)
    macid5 =  tkinter.Label(top, textvariable=macid5_var, width=15, borderwidth=5).grid(row=7, column=0, padx=10,
                                                                                       pady=10)
    macid6 =  tkinter.Label(top, textvariable=macid6_var, width=15, borderwidth=5).grid(row=8, column=0, padx=10,
                                                                                       pady=10)
    macid7 = tkinter.Label(top, textvariable=macid7_var, width=15, borderwidth=5).grid(row=9, column=0, padx=10,
                                                                                       pady=10)
    macid8 =  tkinter.Label(top, textvariable=macid8_var, width=15, borderwidth=5).grid(row=10, column=0, padx=10,
                                                                                       pady=10)
    macid9 =  tkinter.Label(top, textvariable=macid9_var, width=15, borderwidth=5).grid(row=11, column=0, padx=10,
                                                                                       pady=10)
    macid10 = tkinter.Label(top, textvariable=macid10_var, width=15, borderwidth=5).grid(row=12, column=0, padx=10,
                                                                                         pady=10)
    macid11 = tkinter.Label(top, textvariable=macid11_var, width=15, borderwidth=5).grid(row=13, column=0, padx=10,
                                                                                         pady=10)
    macid12 =  tkinter.Label(top, textvariable=macid12_var, width=15, borderwidth=5).grid(row=14, column=0, padx=10,
                                                                                         pady=10)

    top.mainloop()


global vm


# Press the green button in the gutter to run the script.
def com_connect(*args):
    vt = clicked.get()
    # va=''
    for nim in ports:
        print(nim)
        if (str(vt) == str(nim)):
            print("hi")
            va = nim.name
            print(va)
    vm = va
    print(vm)
    # print(vt)

def step2():
    root4 = Tk()

    # set the window title
    root4.title('My Window')

    # set the window size
    window_width = 350
    window_height = 270
    screen_width = root4.winfo_screenwidth()
    screen_height = root4.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    root4.geometry('{}x{}+{}+{}'.format(window_width, window_height, x, y))

    def save_text():
        # open the workbook and select the Sheet1
        wb = openpyxl.load_workbook('testresult.xlsx')
        ws = wb['Board1']

        # find the first empty row in the sheet
        row = ws.max_row + 1

        # get the text from the text field and save it in columns M, N, and O
        team = entry.get()
        person = entry1.get()
        designation = entry2.get()
        approved = entry3.get()

        # check if the text fields are empty
        if not all([team, person, designation, approved]):
            messagebox.showwarning('Warning', 'Please fill all the fields.')
            return

        ws.cell(row=row, column=15).value = team
        ws.cell(row=row, column=16).value = person
        ws.cell(row=row, column=17).value = designation
        ws.cell(row=row, column=18).value = approved

        # save the workbook
        wb.save('testresult.xlsx')
        root4.destroy()
        step4()
        # clear the text fields
        entry.delete(0, END)
        entry1.delete(0, END)
        entry2.delete(0, END)
        entry3.delete(0, END)

    # create the GUI


    # create the label and text field
    label = Label(root4, text='Team Name:')

    label.grid(row=0, column=7, padx=10, pady=10)
    entry = Entry(root4)
    entry.grid(row=0, column=8, padx=10, pady=10)
    label = Label(root4, text='Test Person:')

    label.grid(row=1, column=7, padx=10, pady=10)
    entry1 = Entry(root4)
    entry1.grid(row=1, column=8, padx=10, pady=10)
    label = Label(root4, text='Designation:')

    label.grid(row=2, column=7, padx=10, pady=10)
    entry2 = Entry(root4)
    entry2.grid(row=2, column=8, padx=10, pady=10)
    label = Label(root4, text='Approved By:')
    label.grid(row=3, column=7, padx=10, pady=10)
    entry3 = Entry(root4)
    entry3.grid(row=3, column=8, padx=10, pady=10)
    # create the save button

    # create the save button
    button = Button(root4,width=20, pady=7, bg='#57a1f8', fg='white',text='Save', command=save_text)
    button.grid(row=6, column=8, padx=10, pady=10)
    #
    # def next():
    #
    #
    #     root4.destroy()
    #     save_text()
    #
    #     dashbord()
    # Button(root4, width=39, pady=7, text='Next', bg='#57a1f8', fg='white', border=0, command=next).place(x=600, y=450)

    root4.mainloop()

def step4():
    root5 = Tk()

    # set the window title
    root5.title('MAC ID')

    # set the window size
    window_width = 350
    window_height = 270
    screen_width = root5.winfo_screenwidth()
    screen_height = root5.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    root5.geometry('{}x{}+{}+{}'.format(window_width, window_height, x, y))
    # frame2 = Frame(root2, width=650, height=750, bg="white")
    # frame2.place(x=10, y=0)
    level = Label(root5, text='Scan your MAC ID & Verify in Excel', fg='black', wraplength=200,
                  font=('Times New Roman', 14, 'bold'))
    level.place(x=20, y=30)

    # def open_excel():
    #     os.startfile('testresult.xlsx')  # Replace 'filename.xlsx' with the actual name of your Excel file

    def open_excel():
        filename = 'testresult.xlsx'  # Replace 'filename.xlsx' with the actual name of your Excel file
        sheetname = 'MACID'  # Replace 'Sheet2' with the name of the sheet you want to open
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook[sheetname]
        workbook.active = worksheet
        workbook.save(filename)
        os.startfile(filename)

    Button(root5,width=15,pady=5, text='Open Excel', bg='#57a1f8', fg='white', command=open_excel).place(x=220, y=50)

    def next():
        root5.destroy()
        os.system('taskkill /f /im excel.exe')
        dashbord()

    Button(root5, width=18, pady=5, text='Next', bg='#57a1f8', fg='white', border=0, command=next).place(x=210,
                                                                                                         y=200)

    root5.mainloop()


def step3():

    root3 = Toplevel()
    root3.title('Bit stream')
    root3.geometry('925x500+300+200')
    root3.configure(bg="#fff")
    root3.resizable(False, False)
    level1 = Label(root3,
                   text='Run the tool Metha System Updater as Administrator (right mouse button on the icon Metha System Updater and se-lecting the option "Run as Ad-ministrator"  ',
                   fg='#57a1f8', bg='white',
                   font=('Times New Roman', 14, ), wraplength=750)
    level1.place(x=100, y=35)
    level1 = Label(root3,
                   text='STEP1-',
                   fg='black', bg='white',justify='left',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=95, y=105)
    level1 = Label(root3,
                   text='Push the button named ”Con-nect” , after wait several second for verify the Metha S/N and he version ofbitstream installed (FW ver-sion 03.00.05 & HW version 05.04.05). In the example showed 4 metha are connect-ed to the PC used."  ',
                   fg='black', bg='white', justify='left',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=105)

    level1 = Label(root3,
                   text='STEP2-"  ',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=95, y=150)
    level1 = Label(root3,
                   text='Execute the loading of bit-stream selected the icon "fold-er", open of the folder bitstream concerned "  ',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=150)

    level1 = Label(root3,
                   text='STEP3-"  ',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=95, y=175)
    level1 = Label(root3,
                   text='Select the file named “down-load_HW_545_FW_305.bit”"  ',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=175)

    level1 = Label(root3,
                   text='STEP4-',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=95, y=200)
    level1 = Label(root3,
                   text='Select in the menu “Feature” the option “50MHz SDR/DDR VccQ=1.8V – Vcc=3.3V” "  ',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=200)

    level1 = Label(root3,
                   text='STEP5-',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=95, y=225)
    level1 = Label(root3,
                   text='Click first on the "star" icon and after on the icon with the "green arrow" to start the loading of the file "  ',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=225)

    level1 = Label(root3,
                   text='STEP6-',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=95, y=250)
    level1 = Label(root3,
                   text='When the loading was completed the bar under column Result to be green "  ',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=250)

    level1 = Label(root3,
                   text='STEP7-',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=95, y=275)
    level1 = Label(root3,
                   text='Select in the menu “Feature” the option “50MHz SDR/DDR VccQ=3.3V – Vcc=3.3V” and after on the icon with the "green arrow" to start the loading of the file '
                        'At the end reselect file named “download_HW_545_FW_305.bit” and from menu “Feature” the option “50MHz SDR/DDR VccQ=1.8V – Vcc=3.3V” and click the "star" icon."  ',
                   fg='black', bg='white',justify='left',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=275)

    level1 = Label(root3,
                   text='STEP8-',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=95, y=335)
    level1 = Label(root3,
                   text='Once you have prepared the metha by plugging the dut into the daughter and connecting it to the metha and then switching it on, go to the source folder:"  ',
                   fg='black', bg='white',justify='left',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=335)

    level1 = Label(root3,
                   text='STEP9-',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=95, y=375)
    level1 = Label(root3,
                   text='In the source folder at the bottom you will find the .bat files"',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=375)

    level1 = Label(root3,
                   text='STEP10-',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, 'bold'), wraplength=750)
    level1.place(x=90, y=400)
    level1 = Label(root3,
                   text='Then save the file and run it."',
                   fg='black', bg='white',
                   font=('Times New Roman', 12, ), wraplength=750)
    level1.place(x=150, y=400)

    img5 = PhotoImage(file='Msystem.png')

    Label(root3, image=img5, bg='white').place(x=0, y=0)
    root3.mainloop()

def step1():

    root2 = Toplevel()
    root2.title('step1')
    root2.geometry('925x500+300+200')
    root2.configure(bg="#fff")
    root2.resizable(False, False)
    #frame2 = Frame(root2, width=650, height=750, bg="white")
    #frame2.place(x=10, y=0)
    level = Label(root2, text='Fuse Control', fg='#57a1f8', bg='white', font=('Microsoft yaHei UI Light', 20, 'bold'))
    level.place(x=250, y=50)
    level1 = Label(root2,text='On the ASD580A ) check the fuse F1 circuit (Fig.1), with especially must be measured the resist-ing value that must be <1Ω  ',
                   fg='#57a1f8', bg='white',
                   font=('Microsoft yaHei UI Light', 14, 'bold'), wraplength=600)
    level1.place(x=250, y=90)
    level = Label(root2, text='Control of supply voltages on the Motherboard ', fg='#57a1f8', bg='white', font=('Microsoft yaHei UI Light', 20, 'bold'))
    level.place(x=250, y=290)
    level1 = Label(root2,
                   text='Power ON ASD580A  with the power on the bottom showed in the Fig.2. Check with multime-ter the voltage on both side of F1 and verify that the voltage must be ≥12V. ',
                   fg='#57a1f8', bg='white',
                   font=('Microsoft yaHei UI Light', 14, 'bold'), wraplength=680)
    level1.place(x=250, y=330)
    img3 = PhotoImage(file='M3.png')
    Label(root2, image=img3, bg='white', width=200, height=300).place(x=0, y=0)
    img4 = PhotoImage(file='M4.png')
    Label(root2, image=img4, bg='white', width=250, height=200).place(x=0, y=280)


    def next():


        root2.destroy()
        step2()
    Button(root2, width=39, pady=7, text='Next', bg='#57a1f8', fg='white', border=0, command=next).place(x=600, y=450)

    root2.mainloop()



if __name__ == '__main__':
    root1 = Tk()
    root1.title('Login')
    root1.geometry('925x500+300+200')
    root1.configure(bg="#fff")
    root1.resizable(False, False)


    def signin():
        username = user.get()
        password = code.get()

        if username == 'admin' and password == '1234':
            root1.withdraw()
            step1()
            #dashbord()
        elif username != 'admin' and password != '1234':
            messagebox.showerror("Invalid", "invalid username and password")

        elif password != "1234":
            messagebox.showerror("Invalid", "invalid password")

        elif username != 'admin':
            messagebox.showerror("Invalid", "invalid username")


    img = PhotoImage(file='login.png')
    Label(root1, image=img, bg='white').place(x=50, y=50)

    frame = Frame(root1, width=350, height=350, bg="white")
    frame.place(x=480, y=70)

    heading = Label(frame, text='Sign in', fg='#57a1f8', bg='white', font=('Microsoft yaHei UI Light', 23, 'bold'))
    heading.place(x=100, y=5)


    ##############----------------------------------
    def on_enter(e):
        user.delete(0, 'end')


    def on_leave(e):
        name = user.get()
        if name == '':
            user.insert(0, 'Username')


    user = Entry(frame, width=25, fg='black', border=0, bg="white", font=('Microsoft yaHei UI Light', 11, 'bold'))
    user.place(x=30, y=80)
    user.insert(0, 'Username')
    user.bind('<FocusIn>', on_enter)
    user.bind('<FocusOut>', on_leave)

    Frame(frame, width=295, height=2, bg='black').place(x=25, y=107)


    ##############----------------------------------
    def on_enter(event):
        if code.get() == "Password":
            code.delete(0, END)
            code.config(fg='black', show="*")


    def on_leave(event):
        if code.get() == "":
            code.insert(0, 'Password')
            code.config(fg='grey')
        else:
            code.config(show="*")


    def toggle_password():
        if code.cget("show") == "":
            code.config(show="*")
            show_label.config(text="Show")
        else:
            code.config(show="")
            show_label.config(text="Hide")


    code = Entry(frame, width=25, fg='black', border=0, bg="white", font=('Microsoft yaHei UI Light', 11, 'bold'))
    code.place(x=30, y=150)
    code.insert(0, 'Password')
    code.bind('<FocusIn>', on_enter)
    code.bind('<FocusOut>', on_leave)

    show_label = Label(frame, text="Show", fg="black", cursor="hand2")
    show_label.place(x=280, y=150)
    show_label.bind("<Button-1>", lambda event: toggle_password())

    Frame(frame, width=295, height=2, bg='black').place(x=25, y=177)

    ######################################################################

    Button(frame, width=39, pady=7, text='Sign in', bg='#57a1f8', fg='white', border=0, command=signin).place(x=35,
                                                                                                              y=204)

    root1.mainloop()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/


